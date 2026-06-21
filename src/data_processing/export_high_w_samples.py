from __future__ import annotations

import argparse
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


MODULE_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG_PATH = MODULE_DIR / "high_w_sample_export.config.yaml"
SPACE_TO_W = {"X-space": "Wx", "Y-space": "Wy", "Z-space": "Wz"}
W_TO_SPACE = {value: key for key, value in SPACE_TO_W.items()}
GROUP_COLUMNS = ["case_key", "case_label", "method", "fold_id", "source_split_dir", "space"]
MATCH_KEY_PRIORITY = ["ID", "__row_id__", "__source_index__"]
APPENDED_COLUMNS_BASE = [
    "set_type",
    "case_key",
    "method",
    "fold_id",
    "split_id",
    "source_split_dir",
]
APPENDED_COLUMNS_TAIL = ["w_rank_desc", "is_top_n", "is_above_threshold", "highlight_reason"]


@dataclass(frozen=True)
class ExportConfig:
    sample_w_values: Path
    split_summary: Path
    output_root: Path
    workbook_name: str
    overwrite: bool
    top_n: int
    include_all_splits: bool
    thresholds: dict[str, float]
    highlight_test_rows: bool
    highlight_selected_rows: bool
    freeze_header: bool
    add_filter: bool


@dataclass(frozen=True)
class SplitGroup:
    wx_name: str
    space: str
    case_key: str
    case_label: str
    method: str
    fold_id: str
    source_split_dir: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export complete train/test tables for high sample-level W rows, "
            "organized as Wx/Wy/Wz -> case -> method -> fold."
        )
    )
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="YAML config path.")
    parser.add_argument("--dry-run", action="store_true", help="Preview output groups without writing xlsx files.")
    return parser.parse_args()


def resolve_path(raw_path: Any, base_dir: Path) -> Path:
    path = Path(str(raw_path))
    if path.is_absolute():
        return path
    return (base_dir / path).resolve()


def load_config(path: Path) -> ExportConfig:
    if not path.exists():
        raise FileNotFoundError(f"Missing high-W export config: {path}")
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    repo_root = MODULE_DIR.parent.parent
    inputs = payload.get("inputs", {})
    output = payload.get("output", {})
    selection = payload.get("selection", {})
    formatting = payload.get("formatting", {})
    thresholds = selection.get("thresholds", {})
    parsed_thresholds = {
        wx_name: float(thresholds.get(wx_name, float("nan")))
        for wx_name in ["Wx", "Wy", "Wz"]
    }
    for wx_name, threshold in parsed_thresholds.items():
        if not math.isfinite(threshold):
            raise ValueError(f"Missing finite threshold for {wx_name} in {path}")
    return ExportConfig(
        sample_w_values=resolve_path(inputs.get("sample_w_values"), repo_root),
        split_summary=resolve_path(inputs.get("split_summary"), repo_root),
        output_root=resolve_path(output.get("root_dir"), repo_root),
        workbook_name=str(output.get("workbook_name", "data.xlsx")),
        overwrite=bool(output.get("overwrite", True)),
        top_n=int(selection.get("top_n", 10)),
        include_all_splits=bool(selection.get("include_all_splits", True)),
        thresholds=parsed_thresholds,
        highlight_test_rows=bool(formatting.get("highlight_test_rows", True)),
        highlight_selected_rows=bool(formatting.get("highlight_selected_rows", True)),
        freeze_header=bool(formatting.get("freeze_header", True)),
        add_filter=bool(formatting.get("add_filter", True)),
    )


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing input CSV: {path}")
    return pd.read_csv(path, encoding="utf-8-sig")


def validate_inputs(sample_w: pd.DataFrame, split_summary: pd.DataFrame) -> None:
    required_sample = {
        "case_key",
        "case_label",
        "method",
        "fold_id",
        "source_split_dir",
        "space",
        "sample_w_contribution",
    }
    missing_sample = sorted(required_sample - set(sample_w.columns))
    if missing_sample:
        raise ValueError(f"sample_w_values is missing required columns: {missing_sample}")
    required_split = {"source_split_dir", "split_id", "fold_id", "train_file", "test_file"}
    missing_split = sorted(required_split - set(split_summary.columns))
    if missing_split:
        raise ValueError(f"split_summary is missing required columns: {missing_split}")


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def safe_path_part(value: Any) -> str:
    text = normalize_text(value).strip()
    text = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", text, flags=re.UNICODE)
    text = text.strip("._")
    return text or "unknown"


def normalize_source_dir(value: Any) -> str:
    return str(Path(str(value)))


def prepare_sample_w(sample_w: pd.DataFrame) -> pd.DataFrame:
    work = sample_w.copy()
    work = work[work["space"].isin(SPACE_TO_W)].copy()
    work["wx_name"] = work["space"].map(SPACE_TO_W)
    work["sample_w_contribution"] = pd.to_numeric(work["sample_w_contribution"], errors="coerce")
    work = work[work["sample_w_contribution"].notna()].copy()
    work["fold_id"] = work["fold_id"].fillna("").astype(str)
    work["source_split_dir_norm"] = work["source_split_dir"].map(normalize_source_dir)
    return work


def prepare_split_summary(split_summary: pd.DataFrame) -> pd.DataFrame:
    work = split_summary.copy()
    work["source_split_dir_norm"] = work["source_split_dir"].map(normalize_source_dir)
    work["fold_id"] = work["fold_id"].fillna("").astype(str)
    return work


def build_split_lookup(split_summary: pd.DataFrame) -> dict[str, pd.Series]:
    lookup: dict[str, pd.Series] = {}
    for _, row in split_summary.iterrows():
        key = str(row["source_split_dir_norm"])
        lookup.setdefault(key, row)
    return lookup


def iter_groups(sample_w: pd.DataFrame) -> Iterable[tuple[SplitGroup, pd.DataFrame]]:
    sort_cols = ["wx_name", "case_key", "method", "fold_id", "source_split_dir_norm"]
    work = sample_w.sort_values(sort_cols, kind="stable")
    for keys, group in work.groupby(["wx_name", *GROUP_COLUMNS], dropna=False, sort=False):
        wx_name, case_key, case_label, method, fold_id, source_split_dir, space = keys
        split_group = SplitGroup(
            wx_name=str(wx_name),
            space=str(space),
            case_key=str(case_key),
            case_label=str(case_label),
            method=str(method),
            fold_id=normalize_text(fold_id),
            source_split_dir=str(source_split_dir),
        )
        yield split_group, group.copy()


def output_path_for_group(config: ExportConfig, group: SplitGroup) -> Path:
    base = config.output_root / group.wx_name / safe_path_part(group.case_key) / safe_path_part(group.method)
    if group.fold_id:
        base = base / safe_path_part(group.fold_id)
    return base / config.workbook_name


def choose_match_key(test_df: pd.DataFrame, w_rows: pd.DataFrame) -> str | None:
    for key in MATCH_KEY_PRIORITY:
        if key not in test_df.columns or key not in w_rows.columns:
            continue
        test_keys = pd.to_numeric(test_df[key], errors="coerce")
        w_keys = pd.to_numeric(w_rows[key], errors="coerce")
        if test_keys.notna().all() and w_keys.notna().all() and test_keys.is_unique and w_keys.is_unique:
            return key
    return None


def build_w_lookup(w_rows: pd.DataFrame, match_key: str, top_n: int, threshold: float) -> dict[Any, dict[str, Any]]:
    rows = w_rows.copy()
    rows = rows.dropna(subset=["sample_w_contribution"])
    rows = rows.sort_values("sample_w_contribution", ascending=False, kind="stable").reset_index(drop=True)
    rows["computed_rank"] = rows.index + 1
    lookup: dict[Any, dict[str, Any]] = {}
    for _, row in rows.iterrows():
        raw_key = row[match_key]
        key = pd.to_numeric(pd.Series([raw_key]), errors="coerce").iloc[0]
        if pd.isna(key):
            continue
        w_value = float(row["sample_w_contribution"])
        rank_value = int(row["computed_rank"])
        is_top_n = rank_value <= top_n
        is_above_threshold = w_value >= threshold
        reasons: list[str] = []
        if is_top_n:
            reasons.append(f"top_{top_n}")
        if is_above_threshold:
            reasons.append("above_threshold")
        lookup[key] = {
            "w_value": w_value,
            "w_rank_desc": rank_value,
            "is_top_n": is_top_n,
            "is_above_threshold": is_above_threshold,
            "highlight_reason": ";".join(reasons),
        }
    return lookup


def append_metadata(
    frame: pd.DataFrame,
    *,
    set_type: str,
    group: SplitGroup,
    split_row: pd.Series,
    wx_name: str,
) -> pd.DataFrame:
    result = frame.copy()
    result["set_type"] = set_type
    result["case_key"] = group.case_key
    result["method"] = group.method
    result["fold_id"] = group.fold_id
    result["split_id"] = split_row.get("split_id", "")
    result["source_split_dir"] = group.source_split_dir
    result[wx_name] = pd.NA
    result["w_rank_desc"] = pd.NA
    result["is_top_n"] = False
    result["is_above_threshold"] = False
    result["highlight_reason"] = ""
    return result


def attach_test_w_values(test_df: pd.DataFrame, lookup: dict[Any, dict[str, Any]], match_key: str, wx_name: str) -> pd.DataFrame:
    result = test_df.copy()
    for idx, raw_key in result[match_key].items():
        key = pd.to_numeric(pd.Series([raw_key]), errors="coerce").iloc[0]
        payload = lookup.get(key)
        if payload is None:
            continue
        result.at[idx, wx_name] = payload["w_value"]
        result.at[idx, "w_rank_desc"] = payload["w_rank_desc"]
        result.at[idx, "is_top_n"] = bool(payload["is_top_n"])
        result.at[idx, "is_above_threshold"] = bool(payload["is_above_threshold"])
        result.at[idx, "highlight_reason"] = payload["highlight_reason"]
    return result


def ordered_columns(original_columns: list[str], wx_name: str) -> list[str]:
    appended = [*APPENDED_COLUMNS_BASE, wx_name, *APPENDED_COLUMNS_TAIL]
    return [column for column in original_columns if column not in appended] + appended


def normalize_excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    return value


def estimate_column_width(values: list[Any], header: str) -> float:
    strings = [str(header), *[str(value) for value in values[:250] if value is not None]]
    max_len = max((len(text) for text in strings), default=len(header))
    return float(min(max(max_len + 2, 10), 42))


def write_workbook(path: Path, data: pd.DataFrame, wx_name: str, config: ExportConfig) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "data"

    headers = list(data.columns)
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    test_fill = PatternFill("solid", fgColor="EAF3F8")
    selected_fill = PatternFill("solid", fgColor="FFD966")
    threshold_fill = PatternFill("solid", fgColor="F4B183")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    set_type_col = headers.index("set_type") + 1
    top_n_col = headers.index("is_top_n") + 1
    threshold_col = headers.index("is_above_threshold") + 1
    wx_col = headers.index(wx_name) + 1

    for _, row in data.iterrows():
        values = [normalize_excel_value(row[column]) for column in headers]
        sheet.append(values)
        excel_row = sheet.max_row
        is_test = row["set_type"] == "test"
        is_top_n = bool(row["is_top_n"])
        is_above_threshold = bool(row["is_above_threshold"])
        if config.highlight_selected_rows and is_above_threshold:
            fill = threshold_fill
        elif config.highlight_selected_rows and is_top_n:
            fill = selected_fill
        elif config.highlight_test_rows and is_test:
            fill = test_fill
        else:
            fill = None
        if fill is not None:
            for cell in sheet[excel_row]:
                cell.fill = fill

    if config.freeze_header:
        sheet.freeze_panes = "A2"
    if config.add_filter:
        sheet.auto_filter.ref = sheet.dimensions

    for column_index, header in enumerate(headers, start=1):
        column_values = [normalize_excel_value(value) for value in data[header].tolist()]
        sheet.column_dimensions[get_column_letter(column_index)].width = estimate_column_width(column_values, header)
    sheet.column_dimensions[get_column_letter(set_type_col)].width = 12
    sheet.column_dimensions[get_column_letter(wx_col)].width = 14
    sheet.column_dimensions[get_column_letter(top_n_col)].width = 12
    sheet.column_dimensions[get_column_letter(threshold_col)].width = 18

    workbook.save(path)


def build_export_frame(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    *,
    group: SplitGroup,
    split_row: pd.Series,
    w_rows: pd.DataFrame,
    match_key: str,
    config: ExportConfig,
    threshold: float,
) -> tuple[pd.DataFrame, int, int, int]:
    wx_name = group.wx_name
    lookup = build_w_lookup(w_rows, match_key, config.top_n, threshold)
    train_meta = append_metadata(train_df, set_type="train", group=group, split_row=split_row, wx_name=wx_name)
    test_meta = append_metadata(test_df, set_type="test", group=group, split_row=split_row, wx_name=wx_name)
    test_meta = attach_test_w_values(test_meta, lookup, match_key, wx_name)
    original_columns = list(train_df.columns)
    columns = ordered_columns(original_columns, wx_name)
    combined = pd.concat([train_meta[columns], test_meta[columns]], ignore_index=True)
    selected_count = int((test_meta["is_top_n"] | test_meta["is_above_threshold"]).sum())
    threshold_count = int(test_meta["is_above_threshold"].sum())
    top_n_count = int(test_meta["is_top_n"].sum())
    return combined, selected_count, threshold_count, top_n_count


def manifest_record(
    *,
    group: SplitGroup,
    output_path: Path,
    threshold: float,
    top_n: int,
    status: str,
    message: str = "",
    selected_count: int = 0,
    threshold_count: int = 0,
    top_n_count: int = 0,
    train_rows: int = 0,
    test_rows: int = 0,
    match_key: str = "",
) -> dict[str, Any]:
    return {
        "space": group.wx_name,
        "space_label": group.space,
        "case_key": group.case_key,
        "case_label": group.case_label,
        "method": group.method,
        "fold_id": group.fold_id,
        "source_split_dir": group.source_split_dir,
        "output_path": str(output_path),
        "threshold": threshold,
        "top_n": top_n,
        "selected_count": selected_count,
        "threshold_count": threshold_count,
        "top_n_count": top_n_count,
        "train_rows": train_rows,
        "test_rows": test_rows,
        "match_key": match_key,
        "status": status,
        "message": message,
    }


def export_groups(config: ExportConfig, dry_run: bool = False) -> pd.DataFrame:
    sample_w = prepare_sample_w(read_csv(config.sample_w_values))
    split_summary = prepare_split_summary(read_csv(config.split_summary))
    validate_inputs(sample_w, split_summary)
    split_lookup = build_split_lookup(split_summary)
    manifest_rows: list[dict[str, Any]] = []

    for group, w_rows in iter_groups(sample_w):
        output_path = output_path_for_group(config, group)
        threshold = config.thresholds[group.wx_name]
        split_row = split_lookup.get(normalize_source_dir(group.source_split_dir))
        if split_row is None:
            manifest_rows.append(
                manifest_record(
                    group=group,
                    output_path=output_path,
                    threshold=threshold,
                    top_n=config.top_n,
                    status="missing_split_summary",
                    message="source_split_dir not found in split_summary",
                )
            )
            continue
        train_path = Path(str(split_row["train_file"]))
        test_path = Path(str(split_row["test_file"]))
        if not train_path.exists() or not test_path.exists():
            manifest_rows.append(
                manifest_record(
                    group=group,
                    output_path=output_path,
                    threshold=threshold,
                    top_n=config.top_n,
                    status="missing_split_file",
                    message=f"train/test file missing: {train_path} | {test_path}",
                )
            )
            continue
        train_df = read_csv(train_path)
        test_df = read_csv(test_path)
        match_key = choose_match_key(test_df, w_rows)
        if match_key is None:
            manifest_rows.append(
                manifest_record(
                    group=group,
                    output_path=output_path,
                    threshold=threshold,
                    top_n=config.top_n,
                    status="missing_unique_match_key",
                    message="Could not uniquely align test rows using ID, __row_id__, or __source_index__",
                    train_rows=len(train_df),
                    test_rows=len(test_df),
                )
            )
            continue
        try:
            combined, selected_count, threshold_count, top_n_count = build_export_frame(
                train_df,
                test_df,
                group=group,
                split_row=split_row,
                w_rows=w_rows,
                match_key=match_key,
                config=config,
                threshold=threshold,
            )
        except Exception as error:  # noqa: BLE001 - preserve failure in manifest for batch exports.
            manifest_rows.append(
                manifest_record(
                    group=group,
                    output_path=output_path,
                    threshold=threshold,
                    top_n=config.top_n,
                    status="build_failed",
                    message=str(error),
                    train_rows=len(train_df),
                    test_rows=len(test_df),
                    match_key=match_key,
                )
            )
            continue

        if output_path.exists() and not config.overwrite and not dry_run:
            status = "skipped_exists"
            message = "output exists and overwrite is false"
        else:
            status = "dry_run" if dry_run else "ok"
            message = ""
            if not dry_run:
                write_workbook(output_path, combined, group.wx_name, config)
        manifest_rows.append(
            manifest_record(
                group=group,
                output_path=output_path,
                threshold=threshold,
                top_n=config.top_n,
                status=status,
                message=message,
                selected_count=selected_count,
                threshold_count=threshold_count,
                top_n_count=top_n_count,
                train_rows=len(train_df),
                test_rows=len(test_df),
                match_key=match_key,
            )
        )

    manifest = pd.DataFrame(manifest_rows)
    config.output_root.mkdir(parents=True, exist_ok=True)
    manifest_path = config.output_root / "manifest.csv"
    manifest.to_csv(manifest_path, index=False, encoding="utf-8-sig")
    return manifest


def print_summary(manifest: pd.DataFrame, output_root: Path, dry_run: bool) -> None:
    action = "Dry-run preview" if dry_run else "Export complete"
    print(f"{action}: {output_root}")
    if manifest.empty:
        print("No groups found.")
        return
    by_space = manifest.groupby(["space", "status"], dropna=False).size().reset_index(name="count")
    print(by_space.to_string(index=False))
    failures = manifest[~manifest["status"].isin(["ok", "dry_run", "skipped_exists"])]
    if not failures.empty:
        print("\nFailures:")
        print(failures[["space", "case_key", "method", "fold_id", "status", "message"]].head(20).to_string(index=False))
    print(f"Manifest: {output_root / 'manifest.csv'}")


def main() -> None:
    args = parse_args()
    config = load_config(Path(args.config))
    manifest = export_groups(config, dry_run=bool(args.dry_run))
    print_summary(manifest, config.output_root, bool(args.dry_run))


if __name__ == "__main__":
    main()
